from selenium import webdriver
from selenium.webdriver.common.keys import Keys    
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from time import sleep
import openpyxl
import customtkinter as ctk
import tkinter as tk
from PIL import Image, ImageTk

numero_oab = ""
estado = ""

janela = ctk.CTk()

class Funcs(): 
    def backend(self):
        global numero_oab, estado
        chrome_options = Options()
        chrome_options.add_argument("--headless")

        # Entrar no site da https://pje-consulta-publica.tjmg.jus.br/
        driver = webdriver.Chrome(options=chrome_options)
        driver.get('https://pje-consulta-publica.tjmg.jus.br/')
        sleep(3)
        # Digitar número OAB e selecionar estado
        campo_oab = driver.find_element(By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']") #Seleciona o campo onde digitamos o OAB number
        campo_oab.send_keys(numero_oab) #Digitamos o número de numero_oab dentro do campo
        # Selecionar o estado
        dropdown_estados = driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
        opcoes_estados = Select(dropdown_estados)
        opcoes_estados.select_by_visible_text(estado)
        # Clicar em pesquisar
        botao_pesquisar = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
        botao_pesquisar.click()
        sleep(2)
        # Entrar em cada um dos processos
        processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")

        for processo in processos:
            #processo = processos [0]
            processo.click()
            janelas = driver.window_handles #Retorna o código das janelas abertas I34UG... KLIO...
            driver.switch_to.window(janelas[-1]) #Seleciona a janela aberta por último
            driver.set_window_size(1920,1080)

            # Extrair o n° do processo e data da distribuição
            numero_processo = driver.find_elements(By.XPATH, "//div[@class='col-sm-12 ']")
            numero_processo = numero_processo[0] #Seleciona a primeira div encontrada
            numero_processo = numero_processo.text

            data_distribuicao = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")
            data_distribuicao = data_distribuicao[1]
            data_distribuicao = data_distribuicao.text

            # Extrair e guardar todas as últimas movimentações
            movimentacoes = driver.find_elements(By.XPATH,"//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class,'rich-table-row ')]//td//div//div//span")
            lista_movimentacoes = []
            for movimentacao in movimentacoes:
                lista_movimentacoes.append(movimentacao.text)

            # Guardar tudo numa planilha do excel
            workbook = openpyxl.load_workbook('dados.xlsx')
            try:
                #Código para inserir dados em página existente e inserir as informações
                #Acessar página do processos
                pagina_processo = workbook[numero_processo]
                #Criar nome das colunas
                pagina_processo['A1'].value ="Número Processo"
                pagina_processo['B1'].value ="Data distribuição"
                pagina_processo['C1'].value ="Movimentações"
                #adicionar o número do processo
                pagina_processo['A2'].value = numero_processo
                #adicionar data de distribuição
                pagina_processo['B2'].value = data_distribuicao
                #adicionar movimentações
                for index, linha in enumerate(pagina_processo.iter_rows(min_row=2,min_col=3,max_col=3,max_row=len(lista_movimentacoes))):
                    for celula in linha:
                        celula.value = lista_movimentacoes[index]

                workbook.save('dados.xlsx')
                driver.close()
                sleep(1)
                driver.switch_to.window(driver.window_handles[0])
            except Exception as error:
                #Código para inserir dados em página existente e inserir as informações
                workbook.create_sheet(numero_processo)
                #Acessar página do processos
                pagina_processo = workbook[numero_processo]
                #Criar nome das colunas
                pagina_processo['A1'].value ="Número Processo"
                pagina_processo['B1'].value ="Data distribuição"
                pagina_processo['C1'].value ="Movimentações"
                #adicionar o número do processo
                pagina_processo['A2'].value = numero_processo
                #adicionar data de distribuição
                pagina_processo['B2'].value = data_distribuicao
                #adicionar movimentações
                for index, linha in enumerate(pagina_processo.iter_rows(min_row=2,min_col=3,max_col=3,max_row=len(lista_movimentacoes))):
                    for celula in linha:
                        celula.value = lista_movimentacoes[index]

                workbook.save('dados.xlsx')
                driver.close()
                sleep(1)
                driver.switch_to.window(driver.window_handles[0])

class Application(Funcs):
    def __init__(self):
        self.root = janela
        self.tela()
        self.frames_da_tela()
        self.widgets_frame_1()
        self.widgets_frame_2()


        janela.mainloop()
    def tela(self):
        self.root.title("OAB process collector")
        self.root.geometry('600x400')
        self.root.resizable(False,False)

    def frames_da_tela(self):
        #self.frame_1 = ctk.CTkFrame(master=janela, width=340, height=380, fg_color="teal", bg_color="red", border_width=10, corner_radius=30).place(x=10, y=10)
        self.frame_1 = ctk.CTkFrame(janela, width=340, height=380).place(x=10, y=10)
        self.frame_2 = ctk.CTkFrame(janela, width=220 , height=120).place(x=365, y=140)

    def widgets_frame_1(self):
        #ASímbolo da empresa
        self.simb = Image.open("Fontenele.png")
        self.simb = self.simb.resize((300, 200))
        self.imagem_tk = ImageTk.PhotoImage(self.simb)
        self.label_imagem = tk.Label(self.root, image=self.imagem_tk, highlightbackground="black", highlightthickness=1)
        self.label_imagem.place(x= 25, y= 90)

    def widgets_frame_2(self):
        #Texto inicial
        label = ctk.CTkLabel(self.frame_2, text="Search system", font=("Roboto", 20))
        label.place(x=407, y=108)
        #Caixa de texto
        self.entry_1 = ctk.CTkEntry(self.frame_2, placeholder_text="OAB Number", font=("Roboto", 14))
        self.entry_1.place(x=405, y=150)
        self.dropdown = ctk.CTkOptionMenu(self.frame_2, values=["AM", "SP", "AC", "AL", "AP", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SE", "TO"])
        self.dropdown.place(x=405, y=185)
        self.button = ctk.CTkButton(master=self.frame_2, text = "FIND", command=self.search). place(x=405, y=220)

    def search(self):
        global numero_oab, estado
        numero_oab = self.entry_1.get()
        estado = self.dropdown.get()

        self.backend()



Application()
